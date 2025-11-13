"""TODO: add documentation."""

import os
from typing import Any

import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font

from jmeter_test_suite.domain.entities.excel_report import ExcelReport


class ExcelAdapter:
    """TODO: add documentation."""

    def __init__(self) -> None:
        """TODO: add documentation."""

    def generate_excel_report(
        self,
        excel_report: ExcelReport,
        jmeter_data: dict[str, Any],
        nmon_data: dict[str, Any],
    ) -> bool:
        """TODO: add documentation."""
        try:
            # Update report status
            excel_report.status = "processing"

            # Create workbook
            wb = openpyxl.Workbook()

            # Create worksheet
            self._create_summary_sheet(wb, jmeter_data, nmon_data)
            self._create_jmeter_sheet(wb, jmeter_data)
            self._create_nmon_sheet(wb, nmon_data)

            # 如果有图表需求，Create charts
            if excel_report.include_charts:
                self._create_charts(wb, jmeter_data, nmon_data)

            # Create output directory
            output_dir = os.path.dirname(excel_report.output_file)
            if output_dir:
                os.makedirs(output_dir, exist_ok=True)

            # Save Excel file
            wb.save(excel_report.output_file)

            # Update report status
            excel_report.status = "completed"

            return True

        except Exception as e:
            excel_report.status = "failed"
            print(f"生成Excel报告失败: {str(e)}")
            return False

    def _create_summary_sheet(
        self,
        wb: openpyxl.Workbook,
        jmeter_data: dict[str, Any],
        nmon_data: dict[str, Any],
    ) -> None:
        """TODO: add documentation."""
        ws = wb.active
        ws.title = "汇总报告"

        # Set title style
        title_font = Font(name="微软雅黑", size=16, bold=True)
        header_font = Font(name="微软雅黑", size=12, bold=True)
        normal_font = Font(name="微软雅黑", size=10)

        # 标题
        ws["A1"] = "JMeter和nmon性能测试汇总报告"
        ws["A1"].font = title_font
        ws.merge_cells("A1:D1")

        # JMeter数据汇总
        row = 3
        ws[f"A{row}"] = "JMeter性能数据"
        ws[f"A{row}"].font = header_font

        jmeter_summary = [
            ("总样本数", jmeter_data.get("total_samples", 0)),
            ("成功样本数", jmeter_data.get("successful_samples", 0)),
            ("失败样本数", jmeter_data.get("failed_samples", 0)),
            ("平均响应时间(ms)", jmeter_data.get("average_response_time", 0.0)),
            ("TPS", jmeter_data.get("tps", 0.0)),
            ("错误率(%)", jmeter_data.get("error_rate", 0.0)),
        ]

        for item, value in jmeter_summary:
            row += 1
            ws[f"A{row}"] = item
            ws[f"B{row}"] = value
            ws[f"A{row}"].font = normal_font
            ws[f"B{row}"].font = normal_font

        # nmon数据汇总
        row += 2
        ws[f"A{row}"] = "nmon系统监控数据"
        ws[f"A{row}"].font = header_font

        nmon_summary = [
            ("CPU使用率平均值(%)", nmon_data.get("cpu_usage_avg", 0.0)),
            ("内存使用率平均值(%)", nmon_data.get("memory_usage_avg", 0.0)),
            ("磁盘IO平均值(%)", nmon_data.get("disk_io_avg", 0.0)),
            ("网络IO平均值(%)", nmon_data.get("network_io_avg", 0.0)),
        ]

        for item, value in nmon_summary:
            row += 1
            ws[f"A{row}"] = item
            ws[f"B{row}"] = value
            ws[f"A{row}"].font = normal_font
            ws[f"B{row}"].font = normal_font

        # Adjust column width
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15

    def _create_jmeter_sheet(
        self, wb: openpyxl.Workbook, jmeter_data: dict[str, Any]
    ) -> None:
        """TODO: add documentation."""
        ws = wb.create_sheet("JMeter性能数据")

        # Set title style
        header_font = Font(name="微软雅黑", size=12, bold=True)
        normal_font = Font(name="微软雅黑", size=10)

        # Header row
        headers = ["指标", "数值", "单位"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font

        # Data rows
        jmeter_items = [
            ("总样本数", jmeter_data.get("total_samples", 0), "个"),
            ("成功样本数", jmeter_data.get("successful_samples", 0), "个"),
            ("失败样本数", jmeter_data.get("failed_samples", 0), "个"),
            ("平均响应时间", jmeter_data.get("average_response_time", 0.0), "ms"),
            ("TPS", jmeter_data.get("tps", 0.0), "req/s"),
            ("错误率", jmeter_data.get("error_rate", 0.0), "%"),
        ]

        for row, (item, value, unit) in enumerate(jmeter_items, 2):
            ws.cell(row=row, column=1, value=item).font = normal_font
            ws.cell(row=row, column=2, value=value).font = normal_font
            ws.cell(row=row, column=3, value=unit).font = normal_font

        # Adjust column width
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 10

    def _create_nmon_sheet(
        self, wb: openpyxl.Workbook, nmon_data: dict[str, Any]
    ) -> None:
        """TODO: add documentation."""
        ws = wb.create_sheet("nmon系统监控数据")

        # Set title style
        header_font = Font(name="微软雅黑", size=12, bold=True)
        normal_font = Font(name="微软雅黑", size=10)

        # Header row
        headers = ["系统资源", "使用率", "单位"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font

        # Data rows
        nmon_items = [
            ("CPU使用率", nmon_data.get("cpu_usage_avg", 0.0), "%"),
            ("内存使用率", nmon_data.get("memory_usage_avg", 0.0), "%"),
            ("磁盘IO", nmon_data.get("disk_io_avg", 0.0), "%"),
            ("网络IO", nmon_data.get("network_io_avg", 0.0), "%"),
        ]

        for row, (resource, value, unit) in enumerate(nmon_items, 2):
            ws.cell(row=row, column=1, value=resource).font = normal_font
            ws.cell(row=row, column=2, value=value).font = normal_font
            ws.cell(row=row, column=3, value=unit).font = normal_font

        # Adjust column width
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 10

    def _create_charts(
        self,
        wb: openpyxl.Workbook,
        _jmeter_data: dict[str, Any],
        _nmon_data: dict[str, Any],
    ) -> None:
        """TODO: add documentation."""
        try:
            # 在JMeter工作表中Create charts
            jmeter_ws = wb["JMeter性能数据"]

            # CreateTPS柱状图
            chart = BarChart()
            chart.title = "JMeter性能指标"
            chart.x_axis.title = "指标"
            chart.y_axis.title = "数值"

            # 添加数据（简化版本）
            data = Reference(jmeter_ws, min_col=2, min_row=2, max_row=6, max_col=2)
            cats = Reference(jmeter_ws, min_col=1, min_row=2, max_row=6, max_col=1)

            chart.add_data(data, titles_from_data=False)
            chart.set_categories(cats)

            # 将图表添加到工作表
            jmeter_ws.add_chart(chart, "E2")

        except Exception as e:
            print(f"创建图表失败: {str(e)}")
            # 图表Create失败不影响整体功能
