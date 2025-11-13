"""TODO: add documentation."""

from collections.abc import Iterable
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

ParsedJtl = dict[str, Any]
NmonSummary = dict[str, float]
JtlFileIterable = Iterable[tuple[int, str]]


class MultiThreadExcelAdapter:
    """TODO: add documentation."""

    def __init__(self) -> None:
        """TODO: add documentation."""

    def generate_multi_thread_report(
        self, jtl_files: JtlFileIterable, nmon_data_file: str, output_file: str
    ) -> bool:
        """TODO: add documentation."""
        try:
            # Create workbook
            wb = Workbook()
            wb.remove(wb.active)

            # 1. Create多线程测试报告工作表
            ws1 = wb.create_sheet("多线程测试报告", 0)
            self._create_multi_thread_sheet(ws1, jtl_files, nmon_data_file)

            # 2. Create详细数据工作表
            ws2 = wb.create_sheet("详细测试数据", 1)
            self._create_detail_data_sheet(ws2, jtl_files)

            # 保存文件
            wb.save(output_file)
            print(f"✅ 多线程测试报告已生成: {output_file}")
            return True

        except Exception as e:
            print(f"❌ 生成多线程测试报告失败: {str(e)}")
            return False

    def _create_multi_thread_sheet(
        self, ws: Worksheet, jtl_files: JtlFileIterable, nmon_data_file: str
    ) -> None:
        """TODO: add documentation."""

        # 第1行：测试条件
        ws["A1"] = (
            "测试条件：基于真实JMeter测试数据，线程数200、400、600、800，循环数20"
        )
        ws["A1"].font = Font(bold=True)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A1:H1")

        # 第3行：表格标题
        ws["A3"] = "多线程性能测试数据"
        ws["A3"].font = Font(bold=True, color="FFFFFF")
        ws["A3"].fill = PatternFill(
            start_color="FF5A5A5A", end_color="FF5A5A5A", fill_type="solid"
        )
        ws["A3"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A3:H3")

        # 第4行：表头
        headers = [
            "线程数",
            "循环数",
            "总请求数",
            "成功数",
            "失败数",
            "TPS",
            "平均响应时间(ms)",
            "最大响应时间(ms)",
        ]
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=i)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="FF5A5A5A", end_color="FF5A5A5A", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")

        # Parse所有JTL文件
        test_results = []
        for threads, jtl_file in jtl_files:
            result = self._parse_jtl_file(jtl_file)
            if result:
                result["threads"] = threads
                result["loops"] = 20
                test_results.append(result)

        # 写入真实测试数据
        for row_idx, result in enumerate(test_results, 5):
            # 交替行背景色
            bg_color = "FFFFFFFF" if row_idx % 2 == 1 else "FFE6E6E6"

            values: list[Any] = [
                result["threads"],
                result["loops"],
                result["total_samples"],
                result["successful_samples"],
                result["failed_samples"],
                result["tps"],
                result["average_response_time"],
                result["max_response_time"],
            ]

            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.fill = PatternFill(
                    start_color=bg_color, end_color=bg_color, fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="center")

        # Parsenmon数据
        nmon_summary = self._parse_nmon_data(nmon_data_file)
        if test_results:
            best_tps = max(result.get("tps", 0.0) for result in test_results)
            best_response = min(
                result.get("average_response_time", 0.0) for result in test_results
            )
        else:
            best_tps = 0.0
            best_response = 0.0

        # 第12行：环境备注和结论
        conclusion_lines = [
            "环境备注：portal:http://192.168.24.45:8080/index",
            "测试地点：西安 带宽：1000M",
            "测试结果：基于真实多线程测试数据，线程数200-800，循环数20",
            f"性能分析：最佳TPS={best_tps:.2f}，最佳响应时间={best_response:.1f}ms",
            "服务器概况：CPU：12th Gen Intel(R) Core(TM) i7-12700 CPU数量：1 "
            "CPU核心数：12 内存：16G",
            (
                "系统监控：CPUuse率="
                f"{nmon_summary.get('cpu_usage', 0):.2f}%，内存use率="
                f"{nmon_summary.get('memory_usage', 0):.2f}%"
            ),
        ]
        conclusion_text = "\n".join(conclusion_lines)

        ws["A12"] = conclusion_text
        ws["A12"].font = Font(bold=True, color="FFFFFF")
        ws["A12"].fill = PatternFill(
            start_color="FF5A5A5A", end_color="FF5A5A5A", fill_type="solid"
        )
        ws["A12"].alignment = Alignment(
            horizontal="left", vertical="top", wrap_text=True
        )
        ws.merge_cells("A12:H12")

        # Set行高和列宽
        ws.row_dimensions[12].height = 120
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 15

        print("✅ 多线程测试报告工作表创建完成")

    def _create_detail_data_sheet(
        self, ws: Worksheet, jtl_files: JtlFileIterable
    ) -> None:
        """TODO: add documentation."""

        # 标题
        ws["A1"] = "详细测试数据对比"
        ws["A1"].font = Font(size=16, bold=True)
        ws.merge_cells("A1:H1")

        # 表头
        headers = [
            "测试场景",
            "线程数",
            "循环数",
            "总请求数",
            "TPS",
            "平均响应时间(ms)",
            "成功率(%)",
            "备注",
        ]
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=i)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")

        # Parse并写入详细数据
        for row_idx, (threads, jtl_file) in enumerate(jtl_files, 4):
            result = self._parse_jtl_file(jtl_file)
            if result:
                success_rate = (
                    result["successful_samples"] / result["total_samples"]
                ) * 100

                values = [
                    f"线程{threads}循环20",
                    threads,
                    20,
                    result["total_samples"],
                    result["tps"],
                    result["average_response_time"],
                    success_rate,
                    "真实测试数据",
                ]

                for col_idx, value in enumerate(values, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.alignment = Alignment(horizontal="center")

        # Set列宽
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 18

        print("✅ 详细测试数据工作表创建完成")

    def _parse_jtl_file(self, jtl_file: str) -> ParsedJtl | None:
        """TODO: add documentation."""
        try:
            df = pd.read_csv(jtl_file)

            total_samples = len(df)
            successful_samples = (
                len(df[df["success"]])
                if "success" in df.columns
                else len(df[df["responseCode"] == "200"])
            )
            failed_samples = total_samples - successful_samples
            average_response_time = (
                df["elapsed"].mean() if "elapsed" in df.columns else 0
            )
            max_response_time = df["elapsed"].max() if "elapsed" in df.columns else 0

            # CalculateTPS
            if "timeStamp" in df.columns and len(df) > 1:
                duration_seconds = (
                    df["timeStamp"].max() - df["timeStamp"].min()
                ) / 1000
                tps = len(df) / duration_seconds if duration_seconds > 0 else 0
            else:
                tps = 0

            print(f"✅ 解析JTL文件 {jtl_file}: {total_samples}样本, TPS={tps:.2f}")

            return {
                "total_samples": total_samples,
                "successful_samples": successful_samples,
                "failed_samples": failed_samples,
                "average_response_time": average_response_time,
                "max_response_time": max_response_time,
                "tps": tps,
            }
        except Exception as e:
            print(f"⚠️ 解析JTL文件失败 {jtl_file}: {str(e)}")
            return None

    def _parse_nmon_data(self, nmon_data_file: str) -> NmonSummary:
        """TODO: add documentation."""
        try:
            df = pd.read_csv(nmon_data_file)

            cpu_usage = df["cpu_usage"].mean() if "cpu_usage" in df.columns else 0
            memory_usage = (
                df["memory_usage"].mean() if "memory_usage" in df.columns else 0
            )
            disk_io = df["disk_io"].mean() if "disk_io" in df.columns else 0

            print(f"✅ 解析nmon数据: CPU={cpu_usage:.2f}%, 内存={memory_usage:.2f}%")

            return {
                "cpu_usage": cpu_usage,
                "memory_usage": memory_usage,
                "disk_io": disk_io,
            }
        except Exception as e:
            print(f"⚠️ 解析nmon数据失败: {str(e)}")
            return {}
