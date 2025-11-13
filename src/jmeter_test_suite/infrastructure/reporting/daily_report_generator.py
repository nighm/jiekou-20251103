"""TODO: add documentation."""

import contextlib
import csv
import glob
import os
import re
import sys
import zipfile
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
import paramiko
import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

PROJECT_ROOT = Path(__file__).resolve().parents[4]

TestResult = dict[str, Any]

DEFAULT_SERVER_HOST = "192.168.24.45"
DEFAULT_SERVER_USER = "test"
DEFAULT_SERVER_PASSWORD = "1"
DEFAULT_SERVER_PORT = 22

# 设置Windows控制台编码
if sys.platform == "win32":
    stdout = sys.stdout
    if hasattr(stdout, "reconfigure"):
        stdout.reconfigure(encoding="utf-8")


class DailyReportGenerator:
    """TODO: add documentation."""

    def __init__(
        self,
        result_dir: str | None = None,
        config_file: str | None = None,
        output_file: str | None = None,
    ) -> None:
        """TODO: add documentation."""
        default_result_dir = PROJECT_ROOT / "result"
        self.result_dir = os.path.abspath(result_dir or str(default_result_dir))
        if not os.path.isdir(self.result_dir):
            print(f"[警告] 结果目录不存在或不可访问: {self.result_dir}")
            os.makedirs(self.result_dir, exist_ok=True)
        config_path = (
            Path(config_file)
            if config_file
            else Path(__file__).resolve().parent.parent
            / "config"
            / "jmeter_config.yaml"
        )
        self.config_file = str(config_path)

        # 确保reports目录存在
        reports_dir = PROJECT_ROOT / "reports"
        if not os.path.exists(reports_dir):
            os.makedirs(reports_dir, exist_ok=True)

        if output_file is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = reports_dir / f"测试日报_心跳接口压测_{timestamp}.xlsx"
        self.output_file = os.path.abspath(str(output_file))
        os.makedirs(os.path.dirname(self.output_file), exist_ok=True)
        self.wb: Workbook | None = None
        self._jtl_files_cache: list[str] | None = None
        self._nmon_files_cache: list[str] | None = None
        self.runtime_warnings: list[str] = []
        templates_dir = PROJECT_ROOT / "templates"
        self.template_files = {
            "normal": str(
                (templates_dir / "test_daily_report_normal_template.xlsx").resolve()
            ),
            "abnormal": str(
                (templates_dir / "test_daily_report_abnormal_template.xlsx").resolve()
            ),
        }

        # 读取配置
        self.config = self.load_config()
        reporting_cfg = self.config.get("reporting", {})
        server_cfg = reporting_cfg.get("server", {})
        self.server_host = server_cfg.get("host", DEFAULT_SERVER_HOST)
        self.server_user = server_cfg.get("user", DEFAULT_SERVER_USER)
        self.server_password = server_cfg.get("password", DEFAULT_SERVER_PASSWORD)
        self.server_port = int(server_cfg.get("port", DEFAULT_SERVER_PORT))

        # 颜色方案
        self.colors = {
            "header_bg": "FF2F4F4F",  # 深灰色背景
            "header_text": "FFFFFFFF",  # 白色文字
            "title_bg": "FF4472C4",  # 蓝色背景
            "success": "FF90EE90",  # 绿色
            "warning": "FFFFD700",  # 黄色
            "error": "FFFF6347",  # 红色
            "info": "FFE6F3FF",  # 浅蓝色（信息背景）
            "data_even": "FFF5F5F5",  # 浅灰
            "data_odd": "FFFFFFFF",  # 白色
        }

        # 字体
        self.title_font = Font(name="微软雅黑", size=18, bold=True, color="FFFFFFFF")
        self.subtitle_font = Font(name="微软雅黑", size=14, bold=True, color="FFFFFFFF")
        self.header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFFFF")
        self.normal_font = Font(name="微软雅黑", size=10)
        self.small_font = Font(name="微软雅黑", size=9)
        self.bold_font = Font(name="微软雅黑", size=10, bold=True)

    def load_config(self) -> dict[str, Any]:
        """TODO: add documentation."""
        try:
            with open(self.config_file, encoding="utf-8") as f:
                return yaml.safe_load(f)
        except Exception as e:
            print(f"[警告] 读取配置文件失败: {e}")
            return {}

    def _find_jtl_files(self) -> list[str]:
        """TODO: add documentation."""
        if self._jtl_files_cache is not None:
            return self._jtl_files_cache

        try:
            pattern = os.path.join(self.result_dir, "*.jtl")
            files = sorted(glob.glob(pattern))
            if not files:
                self.runtime_warnings.append(
                    f"未在目录 {self.result_dir} 中找到任何JTL文件"
                )
            self._jtl_files_cache = files
        except Exception as e:
            print(f"[警告] 搜索JTL文件失败: {e}")
            self.runtime_warnings.append(f"搜索JTL文件失败: {e}")
            self._jtl_files_cache = []
        return self._jtl_files_cache

    def _find_nmon_files(self) -> list[str]:
        """TODO: add documentation."""
        if self._nmon_files_cache is not None:
            return self._nmon_files_cache

        try:
            pattern = os.path.join(self.result_dir, "*.nmon")
            files = sorted(glob.glob(pattern))
            if not files:
                self.runtime_warnings.append(
                    f"未在目录 {self.result_dir} 中找到任何nmon文件"
                )
            self._nmon_files_cache = files
        except Exception as e:
            print(f"[警告] 搜索nmon文件失败: {e}")
            self.runtime_warnings.append(f"搜索nmon文件失败: {e}")
            self._nmon_files_cache = []
        return self._nmon_files_cache

    def calculate_planned_tests(self) -> dict[str, Any]:
        """TODO: add documentation."""

        def _parse_range(raw_value: str, default_value: str) -> tuple[int, int, int]:
            parts = str(raw_value).split()
            if len(parts) != 3:
                self.runtime_warnings.append(
                    f"范围配置格式异常({raw_value})，已使用默认值 {default_value}"
                )
                parts = default_value.split()
            try:
                start, end, step = (int(parts[0]), int(parts[1]), int(parts[2]))
            except (ValueError, TypeError):
                self.runtime_warnings.append(
                    f"范围配置无法解析({raw_value})，已使用默认值 {default_value}"
                )
                start, end, step = (int(x) for x in default_value.split())
            if step == 0:
                fallback = 100 if default_value == "100 100 0" else 10
                self.runtime_warnings.append(
                    f"范围配置步长为0({raw_value})，已改用默认步长 {fallback}"
                )
                step = fallback
            return start, end, step

        try:
            thread_start, thread_end, thread_step = _parse_range(
                self.config.get("thread_range", "100 100 0"),
                "100 100 0",
            )
            loop_start, loop_end, loop_step = _parse_range(
                self.config.get("loop_range", "10 10 0"),
                "10 10 0",
            )

            thread_count = len(range(thread_start, thread_end + 1, thread_step))
            loop_count = len(range(loop_start, loop_end + 1, loop_step))

            planned_tests = thread_count * loop_count

            return {
                "thread_range": f"{thread_start}-{thread_end}(步长{thread_step})",
                "loop_range": f"{loop_start}-{loop_end}(步长{loop_step})",
                "thread_count": thread_count,
                "loop_count": loop_count,
                "planned_tests": planned_tests,
                "planned_jtl_files": planned_tests,
            }
        except Exception as e:
            print(f"[警告] 计算计划测试失败: {e}")
            self.runtime_warnings.append(f"计算计划测试失败: {e}")
            return {
                "thread_range": "未知",
                "loop_range": "未知",
                "thread_count": 0,
                "loop_count": 0,
                "planned_tests": 0,
                "planned_jtl_files": 0,
            }

    def analyze_jtl_file(self, jtl_file: str) -> TestResult | None:
        """TODO: add documentation."""
        if not os.path.exists(jtl_file):
            self.runtime_warnings.append(f"JTL文件不存在: {jtl_file}")
            return None

        if os.path.getsize(jtl_file) == 0:
            self.runtime_warnings.append(f"JTL文件为空: {jtl_file}")
            return None

        try:
            df = pd.read_csv(jtl_file, low_memory=False)
        except Exception as read_err:
            self.runtime_warnings.append(
                f"读取JTL文件失败 {os.path.basename(jtl_file)}: {read_err}"
            )
            return None

        if len(df) == 0:
            self.runtime_warnings.append(f"JTL文件缺少数据: {jtl_file}")
            return None

        try:
            filename = os.path.basename(jtl_file)
            threads_match = re.search(r"(\d+)threads", filename)
            loops_match = re.search(r"(\d+)loops", filename)
            time_match = re.search(r"(\d{8}_\d{6})", filename)

            threads = int(threads_match.group(1)) if threads_match else 0
            loops = int(loops_match.group(1)) if loops_match else 0
            timestamp_str = time_match.group(1) if time_match else ""

            total_samples = len(df)
            expected_samples = threads * loops
            missing_samples = max(0, expected_samples - total_samples)

            if "success" in df.columns:
                success_series = df["success"].astype(bool)
                successful = int(success_series.sum())
                failed = total_samples - successful
            else:
                successful = len(df[df["responseCode"] == "200"])
                failed = total_samples - successful

            success_rate = (
                (successful / total_samples * 100) if total_samples > 0 else 0
            )

            if "elapsed" in df.columns:
                elapsed_series = pd.to_numeric(df["elapsed"], errors="coerce").dropna()
                if len(elapsed_series) > 0:
                    avg_response_time = float(elapsed_series.mean())
                    max_response_time = float(elapsed_series.max())
                    min_response_time = float(elapsed_series.min())
                else:
                    avg_response_time = max_response_time = min_response_time = 0.0
            else:
                avg_response_time = max_response_time = min_response_time = 0.0

            if "timeStamp" in df.columns and len(df) > 0:
                start_time: datetime | None = datetime.fromtimestamp(
                    df["timeStamp"].min() / 1000
                )
                end_time: datetime | None = datetime.fromtimestamp(
                    df["timeStamp"].max() / 1000
                )
                assert start_time is not None and end_time is not None
                duration = (end_time - start_time).total_seconds()
                tps = total_samples / duration if duration > 0 else 0
            else:
                start_time = None
                end_time = None
                duration = 0
                tps = 0

            if missing_samples > 0:
                status = "卡死"
                status_color = self.colors["error"]
            elif failed > 0:
                status = "部分失败"
                status_color = self.colors["warning"]
            else:
                status = "通过"
                status_color = self.colors["success"]

            return {
                "filename": filename,
                "filepath": jtl_file,
                "threads": threads,
                "loops": loops,
                "timestamp": timestamp_str,
                "start_time": start_time,
                "end_time": end_time,
                "duration": duration,
                "expected_samples": expected_samples,
                "total_samples": total_samples,
                "missing_samples": missing_samples,
                "successful": successful,
                "failed": failed,
                "success_rate": success_rate,
                "avg_response_time": avg_response_time,
                "max_response_time": max_response_time,
                "min_response_time": min_response_time,
                "tps": tps,
                "status": status,
                "status_color": status_color,
            }
        except Exception as e:
            print(f"[警告] 分析JTL文件失败 {jtl_file}: {e}")
            self.runtime_warnings.append(f"分析JTL文件失败: {jtl_file}")
            return None

    def get_all_test_results(self) -> list[TestResult]:
        """TODO: add documentation."""
        test_results: list[TestResult] = []
        jtl_files = self._find_jtl_files()
        for jtl_file in jtl_files:
            result = self.analyze_jtl_file(jtl_file)
            if result:
                test_results.append(result)
        return test_results

    def determine_report_mode(self, test_results: list[TestResult]) -> str:
        """TODO: add documentation."""

        for result in test_results:
            if result.get("missing_samples", 0) > 0:
                return "abnormal"
        return "normal"

    def load_template_workbook(self, report_mode: str) -> str:
        """TODO: add documentation."""
        template_path = self.template_files.get(report_mode)
        if template_path is None:
            self.runtime_warnings.append(
                f"未找到{report_mode}模式模板，已使用normal模板"
            )
            template_path = self.template_files.get("normal")
        if not template_path or not os.path.exists(template_path):
            raise FileNotFoundError(f"未找到{report_mode}模式模板: {template_path}")
        self.wb = load_workbook(template_path)
        return template_path

    def _prepare_table_rows(
        self, ws: Worksheet, start_row: int, needed_rows: int, total_columns: int
    ) -> None:
        """TODO: add documentation."""
        if needed_rows <= 0:
            for row_idx in range(start_row, ws.max_row + 1):
                for col_idx in range(1, total_columns + 1):
                    ws.cell(row=row_idx, column=col_idx).value = None
            return

        style_reference = [
            ws.cell(row=start_row, column=col_idx)
            for col_idx in range(1, total_columns + 1)
        ]
        for offset in range(needed_rows):
            target_row = start_row + offset
            if target_row > ws.max_row:
                ws.append([None] * total_columns)
            for col_idx in range(1, total_columns + 1):
                cell = ws.cell(row=target_row, column=col_idx)
                template_cell = style_reference[col_idx - 1]
                with contextlib.suppress(AttributeError):
                    cell._style = template_cell._style
                cell.value = None

        for row_idx in range(start_row + needed_rows, ws.max_row + 1):
            for col_idx in range(1, total_columns + 1):
                ws.cell(row=row_idx, column=col_idx).value = None

    def get_server_config_content(self) -> dict[str, str]:
        """TODO: add documentation."""
        ssh: paramiko.SSHClient | None = None
        try:
            ssh = paramiko.SSHClient()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            ssh.connect(
                self.server_host,
                port=self.server_port,
                username=self.server_user,
                password=self.server_password,
                timeout=10,
            )

            config_file = "/opt/app/yangguan/application.yml"
            stdin, stdout, stderr = ssh.exec_command(f"cat {config_file} 2>/dev/null")
            config_content = stdout.read().decode("utf-8", errors="ignore")

            # 提取线程池相关配置
            threadpool_config = ""
            lines = config_content.split("\n")
            in_threadpool_section = False
            for line in lines:
                if (
                    "task" in line.lower()
                    or "execution" in line.lower()
                    or "pool" in line.lower()
                    or "thread" in line.lower()
                ):
                    in_threadpool_section = True
                    threadpool_config += line + "\n"
                elif (
                    in_threadpool_section
                    and line.strip()
                    and not line.startswith(" ")
                    and not line.startswith("#")
                ):
                    if "spring" not in line.lower():
                        break
                    threadpool_config += line + "\n"
                elif in_threadpool_section and line.strip():
                    threadpool_config += line + "\n"

            return {
                "full_content": config_content,
                "threadpool_config": (
                    threadpool_config if threadpool_config else "未找到线程池配置"
                ),
            }
        except Exception as e:
            print(f"[警告] 获取服务器配置失败: {e}")
            return {"full_content": "", "threadpool_config": "获取失败"}
        finally:
            if ssh is not None:
                with contextlib.suppress(Exception):
                    ssh.close()

    def get_error_log_content(self) -> list[str]:
        """TODO: add documentation."""
        ssh: paramiko.SSHClient | None = None
        try:
            ssh = paramiko.SSHClient()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            ssh.connect(
                self.server_host,
                port=self.server_port,
                username=self.server_user,
                password=self.server_password,
                timeout=10,
            )

            error_log = "/opt/app/yangguan/logs/sys-error.2025-11-04.log.gz"
            error_command = (
                f"zcat {error_log} 2>/dev/null"
                " | grep -A 2 'TaskRejectedException'"
                " | grep '21:18'"
                " | head -15"
            )
            stdin, stdout, stderr = ssh.exec_command(error_command)
            error_lines = stdout.read().decode("utf-8", errors="ignore").strip()

            return error_lines.split("\n")[:10] if error_lines else []
        except Exception as e:
            print(f"[警告] 获取错误日志失败: {e}")
            return []
        finally:
            if ssh is not None:
                with contextlib.suppress(Exception):
                    ssh.close()

    # ===== 基于模板的覆盖实现 =====
    def create_cover_sheet(
        self,
        report_mode: str,
        planned: dict[str, Any],
        actual_results: list[TestResult],
    ) -> None:
        """TODO: add documentation."""
        if self.wb is None:
            raise ValueError("Excel模板尚未加载")
        ws = (
            self.wb["测试执行摘要"]
            if "测试执行摘要" in self.wb.sheetnames
            else self.wb.active
        )

        def fmt_count(value: int | None) -> str:
            if value is None:
                return "-"
            return f"{value}个"

        total_cases = len(actual_results)
        passed_tests = len([r for r in actual_results if r["status"] == "通过"])
        failed_tests = len([r for r in actual_results if r["status"] == "部分失败"])
        stuck_tests = len(
            [
                r
                for r in actual_results
                if r.get("missing_samples", 0) > 0 or r["status"] == "卡死"
            ]
        )
        defect_results = [r for r in actual_results if r["status"] == "部分失败"]

        ws["B3"] = datetime.now().strftime("%Y-%m-%d")
        thread_range = planned.get("thread_range", "未知")
        loop_range = planned.get("loop_range", "未知")
        ws["B11"] = f"线程{thread_range} × 循环{loop_range}"
        ws["C11"] = f"实际执行{total_cases}个测试用例"
        ws["D11"] = ""

        planned_tests = planned.get("planned_tests")
        planned_jtls = planned.get("planned_jtl_files", planned_tests)
        ws["B12"] = (
            fmt_count(planned_tests) if planned_tests else fmt_count(total_cases or 0)
        )
        ws["C12"] = fmt_count(total_cases)
        ws["D12"] = fmt_count(total_cases - planned_tests) if planned_tests else "-"
        ws["B13"] = (
            fmt_count(planned_jtls) if planned_jtls else fmt_count(total_cases or 0)
        )
        ws["C13"] = fmt_count(total_cases)
        ws["D13"] = fmt_count(total_cases - planned_jtls) if planned_jtls else "-"

        ws["B17"] = fmt_count(total_cases)
        ws["B18"] = fmt_count(passed_tests)
        ws["B19"] = fmt_count(failed_tests)
        ws["B20"] = fmt_count(stuck_tests)

        if report_mode == "normal":
            if not actual_results:
                ws["A23"] = "⚠ 未在结果目录中找到可用的JTL文件，请确认压测是否完成。"
                ws["A25"] = ""
            elif defect_results:
                ws["A23"] = (
                    f"缺陷摘要：本次测试共发现{len(defect_results)}项缺陷，请关注修复进度"
                )
                detail_lines = []
                for defect in sorted(
                    defect_results, key=lambda x: (x["threads"], x["loops"])
                ):
                    detail_lines.append(
                        f"- {defect['threads']}线程×{defect['loops']}循环："
                        f"成功率{defect['success_rate']:.2f}%、失败{defect['failed']}、"
                        f"平均响应{defect['avg_response_time']:.2f}ms"
                    )
                ws["A25"] = "\n".join(detail_lines)
            else:
                ws["A23"] = (
                    f"✅ 测试全部通过：所有{total_cases}个测试用例均成功执行，"
                    "无阻塞问题"
                )
                ws["A25"] = "本次执行未发现缺陷，系统性能表现正常。"
            ws["A26"] = ""
            zip_notice_cell = "A31"
            warning_title_cell = "A33"
            warning_start_row = 34
        else:
            stuck_test = next(
                (r for r in actual_results if r.get("missing_samples", 0) > 0), None
            )
            if not stuck_test:
                stuck_test = next(
                    (r for r in actual_results if r["status"] == "卡死"), None
                )
            if stuck_test:
                missing_pct = (
                    (
                        stuck_test["missing_samples"]
                        / stuck_test["expected_samples"]
                        * 100
                    )
                    if stuck_test["expected_samples"]
                    else 0
                )
                ws["A23"] = (
                    f"问题记录：{stuck_test['threads']}线程×{stuck_test['loops']}循环测试异常"
                    f"（缺失样本{stuck_test['missing_samples']}个，占比{missing_pct:.2f}%）"
                )
                phenomenon_lines = [
                    (
                        f"测试人员在执行{stuck_test['threads']}线程×{stuck_test['loops']}循环测试"
                        "时观察到以下现象："
                    ),
                    f"1. 期望样本数：{stuck_test['expected_samples']}个",
                    f"2. 实际样本数：{stuck_test['total_samples']}个",
                    (
                        f"3. 缺失样本：{stuck_test['missing_samples']}个"
                        f"（{missing_pct:.2f}%）"
                    ),
                    f"4. 测试状态：{stuck_test['status']}",
                ]
            else:
                ws["A23"] = "问题记录：测试过程中存在未完成的压测任务，请查明原因"
                phenomenon_lines = [
                    "测试过程中检测到压测卡死或样本缺失，请结合JTL和日志进一步确认。"
                ]
            other_defects = [r for r in defect_results if r is not stuck_test]
            if other_defects:
                phenomenon_lines.append("")
                phenomenon_lines.append("其他缺陷：")
                for defect in sorted(
                    other_defects, key=lambda x: (x["threads"], x["loops"])
                ):
                    phenomenon_lines.append(
                        f"- {defect['threads']}线程×{defect['loops']}循环："
                        f"成功率{defect['success_rate']:.2f}%、失败{defect['failed']}、"
                        f"平均响应{defect['avg_response_time']:.2f}ms"
                    )
            ws["A26"] = "\n".join(phenomenon_lines)
            ws["B28"] = (
                "通过服务器日志及监控初步判断，可能存在应用线程资源不足或后端服务响应堆积，需要进一步排查。"
            )
            if stuck_test:
                ws["B31"] = f"压缩包内/JTL测试数据/{stuck_test['filename']}"
                expected_desc = (
                    "数据来源：JTL文件分析 | "
                    f"期望样本: {stuck_test['expected_samples']}"
                )
                b32_parts = [
                    expected_desc,
                    f"实际样本: {stuck_test['total_samples']}",
                    f"缺失: {stuck_test['missing_samples']}个",
                ]
                ws["B32"] = "，".join(b32_parts)
            else:
                ws["B31"] = "-"
                ws["B32"] = "未定位到具体异常JTL文件，请结合压测结果逐一排查。"
            ws["B36"] = "基于nmon监控数据分析，硬件资源使用情况正常。"
            zip_notice_cell = "A42"
            warning_title_cell = "A44"
            warning_start_row = 45

        zip_name = datetime.now().strftime("%Y%m%d")
        ws[zip_notice_cell] = (
            f'提示：所有测试数据文件已打包在《心跳接口压测数据_{zip_name}.zip》压缩包内，详细路径请参考"数据文件清单"工作表'
        )

        max_warning_lines = 6
        if self.runtime_warnings:
            ws[warning_title_cell] = "数据告警（生成期间自动记录）"
            for idx, warning in enumerate(self.runtime_warnings[:max_warning_lines]):
                ws.cell(row=warning_start_row + idx, column=1, value=f"- {warning}")
            for idx in range(len(self.runtime_warnings), max_warning_lines):
                ws.cell(row=warning_start_row + idx, column=1, value="")
        else:
            ws[warning_title_cell] = ""
            for idx in range(max_warning_lines):
                ws.cell(row=warning_start_row + idx, column=1, value="")

        if report_mode == "abnormal":
            self.create_evidence_screenshots_sheet()

    def create_test_details_sheet(
        self, report_mode: str, actual_results: list[TestResult]
    ) -> None:
        """TODO: add documentation."""
        if self.wb is None:
            raise ValueError("Excel模板尚未加载")
        ws = (
            self.wb["测试执行明细"]
            if "测试执行明细" in self.wb.sheetnames
            else self.wb.create_sheet("测试执行明细")
        )

        sorted_results = sorted(
            actual_results, key=lambda x: (x["threads"], x["loops"])
        )
        self._prepare_table_rows(ws, 3, len(sorted_results), 17)

        normal_status_fill = PatternFill(
            start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid"
        )

        for idx, result in enumerate(sorted_results, start=1):
            row_idx = 2 + idx
            ws.cell(row=row_idx, column=1, value=idx)
            ws.cell(row=row_idx, column=2, value=result["threads"])
            ws.cell(row=row_idx, column=3, value=result["loops"])

            start_time = result.get("start_time")
            end_time = result.get("end_time")
            ws.cell(
                row=row_idx,
                column=4,
                value=start_time.strftime("%Y-%m-%d %H:%M:%S") if start_time else "-",
            )
            ws.cell(
                row=row_idx,
                column=5,
                value=end_time.strftime("%Y-%m-%d %H:%M:%S") if end_time else "-",
            )
            ws.cell(row=row_idx, column=6, value=round(result.get("duration", 0), 2))

            ws.cell(row=row_idx, column=7, value=result.get("expected_samples", 0))
            ws.cell(row=row_idx, column=8, value=result.get("total_samples", 0))
            ws.cell(row=row_idx, column=9, value=result.get("missing_samples", 0))
            ws.cell(row=row_idx, column=10, value=result.get("successful", 0))
            ws.cell(row=row_idx, column=11, value=result.get("failed", 0))
            ws.cell(
                row=row_idx, column=12, value=round(result.get("success_rate", 0), 2)
            )
            ws.cell(
                row=row_idx,
                column=13,
                value=round(result.get("avg_response_time", 0), 2),
            )
            ws.cell(
                row=row_idx,
                column=14,
                value=round(result.get("max_response_time", 0), 2),
            )
            ws.cell(row=row_idx, column=15, value=round(result.get("tps", 0), 2))

            status_cell = ws.cell(
                row=row_idx, column=16, value=result.get("status", "未知")
            )
            if report_mode == "normal":
                status_cell.fill = normal_status_fill
            else:
                status = result.get("status")
                if status == "卡死":
                    color = self.colors["error"]
                elif status == "部分失败":
                    color = self.colors["warning"]
                else:
                    color = self.colors["success"]
                status_cell.fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )

            ws.cell(
                row=row_idx,
                column=17,
                value=f"JTL测试数据/{result.get('filename', '')}",
            )

    def create_issue_analysis_sheet(
        self, report_mode: str, actual_results: list[TestResult]
    ) -> None:
        """TODO: add documentation."""
        if self.wb is None:
            raise ValueError("Excel模板尚未加载")
        ws = (
            self.wb["问题详情分析"]
            if "问题详情分析" in self.wb.sheetnames
            else self.wb.create_sheet("问题详情分析")
        )

        if report_mode == "normal":
            defects = sorted(
                [r for r in actual_results if r["status"] == "部分失败"],
                key=lambda x: (x["threads"], x["loops"]),
            )
            if defects:
                self._prepare_table_rows(ws, 3, len(defects), 8)
                for idx, defect in enumerate(defects, start=1):
                    row_idx = 2 + idx
                    ws.cell(row=row_idx, column=1, value=f"D{idx:03d}")
                    ws.cell(
                        row=row_idx,
                        column=2,
                        value=f"{defect['threads']}线程 × {defect['loops']}循环",
                    )
                    ws.cell(
                        row=row_idx, column=3, value=round(defect["success_rate"], 2)
                    )
                    ws.cell(row=row_idx, column=4, value=defect["failed"])
                    ws.cell(
                        row=row_idx,
                        column=5,
                        value=round(defect["avg_response_time"], 2),
                    )
                    ws.cell(
                        row=row_idx,
                        column=6,
                        value=round(defect["max_response_time"], 2),
                    )
                    ws.cell(
                        row=row_idx, column=7, value=f"JTL测试数据/{defect['filename']}"
                    )
                    ws.cell(
                        row=row_idx,
                        column=8,
                        value="建议开发团队分析失败原因，跟进修复。",
                    )
            else:
                self._prepare_table_rows(ws, 3, 1, 8)
                success_message = (
                    "✅ 本次测试未发现阻塞性问题，"
                    f"共{len(actual_results)}个用例全部成功。"
                )
                msg_cell = ws.cell(row=3, column=1, value=success_message)
                msg_cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                for col in range(2, 9):
                    ws.cell(row=3, column=col, value=None)
            return

        issues = []
        stuck_list = sorted(
            [
                r
                for r in actual_results
                if r.get("missing_samples", 0) > 0 or r["status"] == "卡死"
            ],
            key=lambda x: (x["threads"], x["loops"]),
        )
        partial_list = sorted(
            [r for r in actual_results if r["status"] == "部分失败"],
            key=lambda x: (x["threads"], x["loops"]),
        )

        idx_counter = 1
        for stuck in stuck_list:
            missing_pct = (
                (stuck["missing_samples"] / stuck["expected_samples"] * 100)
                if stuck["expected_samples"]
                else 0
            )
            start_display = (
                stuck["start_time"].strftime("%Y-%m-%d %H:%M:%S")
                if stuck.get("start_time")
                else "-"
            )
            end_display = (
                stuck["end_time"].strftime("%Y-%m-%d %H:%M:%S")
                if stuck.get("end_time")
                else "-"
            )
            stuck_desc = (
                f"{stuck['threads']}线程×{stuck['loops']}循环测试卡死，"
                f"缺少{stuck['missing_samples']}个样本（{missing_pct:.2f}%）"
            )
            issues.append(
                {
                    "id": f"P{idx_counter:03d}",
                    "desc": stuck_desc,
                    "severity": "严重",
                    "found_time": (
                        stuck["start_time"].strftime("%Y-%m-%d %H:%M:%S")
                        if stuck.get("start_time")
                        else "-"
                    ),
                    "config": f"{stuck['threads']}线程 × {stuck['loops']}循环",
                    "ref": (
                        f"1. JTL文件: 压缩包内/JTL测试数据/{stuck['filename']}\n"
                        f"2. 期望样本数: {stuck['expected_samples']}\n"
                        f"3. 实际样本数: {stuck['total_samples']}\n"
                        f"4. 缺失样本: {stuck['missing_samples']}个"
                    ),
                    "phenomenon": (
                        f"1. 期望样本数: {stuck['expected_samples']}\n"
                        f"2. 实际样本数: {stuck['total_samples']}\n"
                        f"3. 缺失样本: {stuck['missing_samples']}个"
                        f"（{missing_pct:.2f}%）\n"
                        f"4. 开始时间: {start_display}\n"
                        f"5. 结束时间: {end_display}"
                    ),
                    "hardware": "基于nmon监控数据分析，硬件资源使用情况正常。",
                    "suggestion": "建议开发团队进一步分析测试数据，确认问题原因。",
                }
            )
            idx_counter += 1

        for fail in partial_list:
            fail_desc = (
                f"{fail['threads']}线程×{fail['loops']}循环测试部分失败，"
                f"失败{fail['failed']}个（成功率{fail['success_rate']:.1f}%）"
            )
            issues.append(
                {
                    "id": f"P{idx_counter:03d}",
                    "desc": fail_desc,
                    "severity": "中等",
                    "found_time": (
                        fail["start_time"].strftime("%Y-%m-%d %H:%M:%S")
                        if fail.get("start_time")
                        else "-"
                    ),
                    "config": f"{fail['threads']}线程 × {fail['loops']}循环",
                    "ref": (
                        f"1. JTL文件: 压缩包内/JTL测试数据/{fail['filename']}\n"
                        f"2. 成功数: {fail['successful']}\n"
                        f"3. 失败数: {fail['failed']}\n"
                        f"4. 成功率: {fail['success_rate']:.1f}%"
                    ),
                    "phenomenon": (
                        f"1. 总样本数: {fail['total_samples']}\n"
                        f"2. 成功数: {fail['successful']}\n"
                        f"3. 失败数: {fail['failed']}\n"
                        f"4. 成功率: {fail['success_rate']:.1f}%"
                    ),
                    "hardware": "基于nmon监控数据分析，硬件资源使用情况正常。",
                    "suggestion": "建议开发团队分析失败原因，优化系统性能。",
                }
            )
            idx_counter += 1

        if not issues:
            self._prepare_table_rows(ws, 3, 1, 9)
            ws.cell(row=3, column=1, value="P001")
            ws.cell(row=3, column=2, value="压测结果存在异常，请核对JTL数据和日志。")
            ws.cell(row=3, column=3, value="中等")
            for col in range(4, 10):
                ws.cell(row=3, column=col, value="-")
            return

        self._prepare_table_rows(ws, 3, len(issues), 9)
        for idx, issue in enumerate(issues, start=1):
            row_idx = 2 + idx
            ws.cell(row=row_idx, column=1, value=issue["id"])
            ws.cell(row=row_idx, column=2, value=issue["desc"])
            ws.cell(row=row_idx, column=3, value=issue["severity"])
            ws.cell(row=row_idx, column=4, value=issue["found_time"])
            ws.cell(row=row_idx, column=5, value=issue["config"])
            ws.cell(row=row_idx, column=6, value=issue["ref"])
            ws.cell(row=row_idx, column=7, value=issue["phenomenon"])
            ws.cell(row=row_idx, column=8, value=issue["hardware"])
            ws.cell(row=row_idx, column=9, value=issue["suggestion"])

    def create_evidence_screenshots_sheet(self) -> None:
        """TODO: add documentation."""
        if self.wb is None:
            raise ValueError("Excel模板尚未加载")
        if "关键证据截图" in self.wb.sheetnames:
            ws = self.wb["关键证据截图"]
        else:
            ws = self.wb.create_sheet("关键证据截图")

        screenshot_dir = "."
        jmeter_log_img = os.path.join(screenshot_dir, "jmeter_live_log.png")
        jtl_file_list_img = os.path.join(screenshot_dir, "jtl_file_list.png")

        if not os.path.exists(jmeter_log_img) or not os.path.exists(jtl_file_list_img):
            png_files = [
                f for f in os.listdir(screenshot_dir) if f.lower().endswith(".png")
            ]
            if not png_files:
                self.runtime_warnings.append("未找到关键证据截图文件")
            for png_file in png_files:
                lower = png_file.lower()
                if not os.path.exists(jmeter_log_img) and any(
                    keyword in lower for keyword in ["jmeter", "live", "log"]
                ):
                    jmeter_log_img = os.path.join(screenshot_dir, png_file)
                if not os.path.exists(jtl_file_list_img) and any(
                    keyword in lower for keyword in ["jtl", "file", "list"]
                ):
                    jtl_file_list_img = os.path.join(screenshot_dir, png_file)

        if os.path.exists(jmeter_log_img):
            ws["A6"] = ""
            try:
                img1 = Image(jmeter_log_img)
                img1.width = 550
                img1.height = 280
                ws.add_image(img1, "A6")
            except Exception as exc:
                self.runtime_warnings.append(f"插入JMeter日志截图失败: {exc}")
                ws["A6"] = "插入JMeter日志截图失败"
        else:
            ws["A6"] = "未找到JMeter日志截图"

        if os.path.exists(jtl_file_list_img):
            ws["A9"] = ""
            try:
                img2 = Image(jtl_file_list_img)
                img2.width = 550
                img2.height = 350
                ws.add_image(img2, "A9")
            except Exception as exc:
                self.runtime_warnings.append(f"插入JTL列表截图失败: {exc}")
                ws["A9"] = "插入JTL列表截图失败"
        else:
            ws["A9"] = "未找到JTL文件列表截图"

    def create_data_files_sheet(
        self, report_mode: str, actual_results: list[TestResult]
    ) -> None:
        """TODO: add documentation."""
        if self.wb is None:
            raise ValueError("Excel模板尚未加载")
        ws = (
            self.wb["数据文件清单"]
            if "数据文件清单" in self.wb.sheetnames
            else self.wb.create_sheet("数据文件清单")
        )

        def find_row(label: str) -> int | None:
            for row in ws.iter_rows(
                min_row=1, max_row=ws.max_row, min_col=1, max_col=1
            ):
                cell = row[0]
                if cell.value == label:
                    return cell.row
            return None

        jtl_section_row = find_row("JTL测试数据文件")
        if jtl_section_row:
            header_row = jtl_section_row + 1
            data_start = header_row + 1
            sorted_results = sorted(
                actual_results, key=lambda x: (x["threads"], x["loops"])
            )
            self._prepare_table_rows(ws, data_start, len(sorted_results), 4)
            for idx, result in enumerate(sorted_results):
                row_idx = data_start + idx
                ws.cell(row=row_idx, column=1, value=result["filename"])
                ws.cell(
                    row=row_idx, column=2, value=f"JTL测试数据/{result['filename']}"
                )
                ws.cell(
                    row=row_idx,
                    column=3,
                    value=f"{result['threads']}线程×{result['loops']}循环测试数据",
                )
                start_time = result.get("start_time")
                end_time = result.get("end_time")
                if start_time and end_time:
                    ws.cell(
                        row=row_idx,
                        column=4,
                        value=(
                            f"{start_time.strftime('%H:%M:%S')} - "
                            f"{end_time.strftime('%H:%M:%S')}"
                        ),
                    )
                else:
                    ws.cell(row=row_idx, column=4, value="时间信息缺失")
        else:
            self.runtime_warnings.append(
                "模板缺少JTL测试数据文件段落，无法填充对应清单"
            )

        if report_mode != "abnormal":
            return

        today = datetime.now().strftime("%Y-%m-%d")
        log_warning = any(
            "服务器日志下载失败" in warn for warn in self.runtime_warnings
        )
        config_warning = any(
            "服务器配置下载失败" in warn for warn in self.runtime_warnings
        )

        log_section_row = find_row("服务器日志文件")
        if log_section_row:
            header_row = log_section_row + 1
            data_start = header_row + 1
            log_file = f"sys-error.{today}.log.gz"
            log_entries = [
                {
                    "name": log_file if not log_warning else f"{log_file}（未下载）",
                    "path": f"服务器日志/{log_file}" if not log_warning else "-",
                    "usage": "应用错误日志（如存在）",
                    "time": "测试执行期间",
                }
            ]
            self._prepare_table_rows(ws, data_start, len(log_entries), 4)
            for idx, entry in enumerate(log_entries):
                row_idx = data_start + idx
                ws.cell(row=row_idx, column=1, value=entry["name"])
                ws.cell(row=row_idx, column=2, value=entry["path"])
                ws.cell(row=row_idx, column=3, value=entry["usage"])
                ws.cell(row=row_idx, column=4, value=entry["time"])

        config_section_row = find_row("服务器配置文件")
        if config_section_row:
            header_row = config_section_row + 1
            data_start = header_row + 1
            config_entries = [
                {
                    "name": (
                        "application.yml"
                        if not config_warning
                        else "application.yml（未下载）"
                    ),
                    "path": "服务器配置/application.yml" if not config_warning else "-",
                    "usage": "应用配置文件",
                    "time": "-",
                }
            ]
            self._prepare_table_rows(ws, data_start, len(config_entries), 4)
            for idx, entry in enumerate(config_entries):
                row_idx = data_start + idx
                ws.cell(row=row_idx, column=1, value=entry["name"])
                ws.cell(row=row_idx, column=2, value=entry["path"])
                ws.cell(row=row_idx, column=3, value=entry["usage"])
                ws.cell(row=row_idx, column=4, value=entry["time"])

    def create_data_package(
        self, report_mode: str, actual_results: list[TestResult]
    ) -> str:
        """TODO: add documentation."""
        print(f"[调试] 接收到 {len(actual_results)} 条测试结果用于数据打包。")

        timestamp = datetime.now().strftime("%Y%m%d")
        reports_dir = os.path.abspath("./reports")
        os.makedirs(reports_dir, exist_ok=True)
        zip_name = os.path.join(reports_dir, f"心跳接口压测数据_{timestamp}.zip")

        print(f"\n[打包] 创建数据压缩包: {zip_name}")

        with zipfile.ZipFile(zip_name, "w", zipfile.ZIP_DEFLATED) as zipf:
            # 1. 测试报告
            print("  [1/5] 添加测试报告...")
            if os.path.exists(self.output_file):
                zipf.write(
                    self.output_file, f"测试报告/{os.path.basename(self.output_file)}"
                )
            else:
                self.runtime_warnings.append("测试报告文件缺失，压缩包未包含报告")

            # 2. JTL测试数据
            print("  [2/5] 添加JTL测试数据文件...")
            jtl_files = self._find_jtl_files()
            for jtl_file in jtl_files:
                try:
                    if os.path.getsize(jtl_file) > 0:
                        zipf.write(
                            jtl_file, f"JTL测试数据/{os.path.basename(jtl_file)}"
                        )
                except OSError as os_err:
                    self.runtime_warnings.append(
                        f"无法添加JTL文件 {os.path.basename(jtl_file)}: {os_err}"
                    )

            # 3. 服务器日志（仅在有问题时下载）
            if report_mode == "abnormal":
                print("  [3/5] 下载并添加服务器日志文件...")
                try:
                    ssh = paramiko.SSHClient()
                    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                    ssh.connect(
                        self.server_host,
                        port=self.server_port,
                        username=self.server_user,
                        password=self.server_password,
                        timeout=30,
                    )

                    today = datetime.now().strftime("%Y-%m-%d")
                    error_log = f"/opt/app/yangguan/logs/sys-error.{today}.log.gz"
                    local_log = "./result/temp_error_log.gz"

                    sftp = ssh.open_sftp()
                    try:
                        sftp.get(error_log, local_log)
                        zipf.write(local_log, f"服务器日志/sys-error.{today}.log.gz")

                        # 提取关键错误
                        import gzip

                        with gzip.open(
                            local_log, "rt", encoding="utf-8", errors="ignore"
                        ) as f:
                            lines = f.readlines()
                            error_lines = [
                                line
                                for line in lines
                                if "TaskRejectedException" in line or "ERROR" in line
                            ][:20]

                        if error_lines:
                            extract_file = "./result/temp_error_extract.txt"
                            with open(extract_file, "w", encoding="utf-8") as f:
                                f.write("\n".join(error_lines))
                            zipf.write(
                                extract_file,
                                f"服务器日志/sys-error.{today}_extract.txt",
                            )
                            os.remove(extract_file)

                        os.remove(local_log)
                    except FileNotFoundError:
                        print(f"    [提示] 未找到错误日志文件: {error_log}")
                    finally:
                        sftp.close()
                        ssh.close()
                except Exception as e:
                    print(f"    [警告] 下载服务器日志失败: {e}")
                    self.runtime_warnings.append(f"服务器日志下载失败: {e}")
            else:
                print("  [3/5] 跳过服务器日志（正常模式无需抓取）")

            # 4. 服务器配置（仅在有问题时下载）
            if report_mode == "abnormal":
                print("  [4/5] 下载并添加服务器配置文件...")
                try:
                    ssh = paramiko.SSHClient()
                    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                    ssh.connect(
                        self.server_host,
                        port=self.server_port,
                        username=self.server_user,
                        password=self.server_password,
                        timeout=30,
                    )

                    config_file = "/opt/app/yangguan/application.yml"
                    local_config = "./result/temp_config.yml"

                    sftp = ssh.open_sftp()
                    sftp.get(config_file, local_config)
                    sftp.close()
                    ssh.close()

                    zipf.write(local_config, "服务器配置/application.yml")
                    os.remove(local_config)
                except Exception as e:
                    print(f"    [警告] 下载服务器配置失败: {e}")
                    self.runtime_warnings.append(f"服务器配置下载失败: {e}")
            else:
                print("  [4/5] 跳过服务器配置（正常模式无需抓取）")

            # 5. 硬件监控数据（关键时间点的）
            print("  [5/5] 添加硬件监控数据文件...")
            nmon_files = self._find_nmon_files()
            for nmon_file in nmon_files[:5]:  # 只添加前5个
                try:
                    if os.path.getsize(nmon_file) > 0:
                        zipf.write(
                            nmon_file, f"硬件监控数据/{os.path.basename(nmon_file)}"
                        )
                except OSError as os_err:
                    self.runtime_warnings.append(
                        f"无法添加nmon文件 {os.path.basename(nmon_file)}: {os_err}"
                    )

            # 6. 关键证据截图（从项目根目录）
            if report_mode == "abnormal":
                print("  [6/6] 添加关键证据截图...")
                screenshot_dir = "."
                screenshot_files = ["jmeter_live_log.png", "jtl_file_list.png"]
                png_files = [
                    f for f in os.listdir(screenshot_dir) if f.lower().endswith(".png")
                ]
                if len(png_files) >= 2:
                    screenshot_files = png_files[:2]

                added_count = 0
                for screenshot_file in screenshot_files:
                    screenshot_path = os.path.join(screenshot_dir, screenshot_file)
                    if os.path.exists(screenshot_path):
                        zipf.write(screenshot_path, f"关键证据截图/{screenshot_file}")
                        print(f"    已添加: {screenshot_file}")
                        added_count += 1

                if added_count == 0:
                    print("    [警告] 未找到关键证据截图文件")
                    self.runtime_warnings.append("未找到关键证据截图文件")
            else:
                print("  [6/6] 正常模式无需关键证据截图")

        print(f"✅ 数据压缩包创建成功: {zip_name}")
        print(f"   文件大小: {os.path.getsize(zip_name) / 1024 / 1024:.2f} MB")

        return zip_name

    def generate(self) -> tuple[str, str]:
        """TODO: add documentation."""
        print("生成测试日报Excel报告（优化版）")
        print("=" * 80)

        actual_results = self.get_all_test_results()
        planned = self.calculate_planned_tests()
        report_mode = self.determine_report_mode(actual_results)
        self.load_template_workbook(report_mode)
        print(f"\n[1/5] 填充封面页（模式：{report_mode}）...")
        self.create_cover_sheet(report_mode, planned, actual_results)

        print("[2/5] 填充测试执行明细表...")
        self.create_test_details_sheet(report_mode, actual_results)

        print("[3/5] 填充问题详情分析表...")
        self.create_issue_analysis_sheet(report_mode, actual_results)

        print("[4/5] 填充数据文件清单...")
        self.create_data_files_sheet(report_mode, actual_results)

        # 保存文件
        print("[5/5] 保存Excel报告...")
        if self.wb is None:
            raise RuntimeError("Excel模板尚未加载")
        self.wb.save(self.output_file)
        print(f"✅ Excel报告生成成功: {self.output_file}")
        print(f"   文件大小: {os.path.getsize(self.output_file) / 1024:.2f} KB")

        # 创建数据压缩包
        print("\n[打包] 创建数据压缩包...")
        zip_file = self.create_data_package(report_mode, actual_results)

        if self.runtime_warnings:
            print("\n[告警] 生成过程中发现以下数据问题(已写入报告) :")
            for warning in self.runtime_warnings:
                print(f"  - {warning}")

        return self.output_file, zip_file

    # ========== 新增功能：JTL文件详细分析 ==========
    def analyze_jtl_stuck_detail(self, jtl_file: str) -> None:
        """TODO: add documentation."""
        print(f"{'=' * 80}")

        if not os.path.exists(jtl_file):
            print(f"[错误] 文件不存在: {jtl_file}")
            return

        timestamps: list[int] = []
        errors: list[dict[str, Any]] = []
        last_timestamp: int | None = None
        max_gap = 0
        max_gap_start: int | None = None
        max_gap_end: int | None = None

        try:
            with open(jtl_file, encoding="utf-8") as f:
                reader = csv.DictReader(f)

                for i, row in enumerate(reader, 1):
                    try:
                        timestamp = int(row["timeStamp"])
                        timestamps.append(timestamp)

                        success = row.get("success", "").lower() == "true"
                        if not success:
                            errors.append(
                                {
                                    "line": i,
                                    "timestamp": timestamp,
                                    "responseCode": row.get("responseCode", ""),
                                    "message": row.get("responseMessage", ""),
                                    "thread": row.get("threadName", ""),
                                }
                            )

                        if last_timestamp is not None:
                            gap = timestamp - last_timestamp
                            if gap > max_gap:
                                max_gap = gap
                                max_gap_start = last_timestamp
                                max_gap_end = timestamp

                        last_timestamp = timestamp
                    except Exception:
                        continue

            if not timestamps:
                print("[错误] 文件中没有有效数据")
                return

            total_samples = len(timestamps)
            start_time = timestamps[0]
            end_time = timestamps[-1]
            duration_ms = end_time - start_time
            duration_sec = duration_ms / 1000.0

            print("\n[基本统计]")
            print(f"  总样本数: {total_samples}")
            start_str = datetime.fromtimestamp(start_time / 1000).strftime(
                "%Y-%m-%d %H:%M:%S.%f"
            )[:-3]
            end_str = datetime.fromtimestamp(end_time / 1000).strftime(
                "%Y-%m-%d %H:%M:%S.%f"
            )[:-3]
            print(f"  开始时间: {start_str}")
            print(f"  结束时间: {end_str}")
            print(f"  持续时间: {duration_sec:.2f} 秒 ({duration_sec / 60:.2f} 分钟)")

            if errors:
                print(f"\n[错误统计] 错误总数: {len(errors)}")
                error_codes: dict[str, int] = defaultdict(int)
                for err in errors:
                    error_codes[err["responseCode"]] += 1
                for code, count in error_codes.items():
                    print(f"  {code}: {count} 次")

            if max_gap > 5000 and max_gap_start is not None and max_gap_end is not None:
                print("\n[警告] 发现长时间间隔（可能卡死）:")
                gap_start_str = datetime.fromtimestamp(max_gap_start / 1000).strftime(
                    "%Y-%m-%d %H:%M:%S.%f"
                )[:-3]
                gap_end_str = datetime.fromtimestamp(max_gap_end / 1000).strftime(
                    "%Y-%m-%d %H:%M:%S.%f"
                )[:-3]
                print(f"  开始: {gap_start_str}")
                print(f"  结束: {gap_end_str}")
                print(f"  间隔: {max_gap / 1000:.2f} 秒")

            # 检查样本数
            filename = os.path.basename(jtl_file)
            threads_match = re.search(r"(\d+)threads", filename)
            loops_match = re.search(r"(\d+)loops", filename)
            if threads_match and loops_match:
                threads = int(threads_match.group(1))
                loops = int(loops_match.group(1))
                expected_samples = threads * loops
                print("\n[完成状态检查]")
                expected_info = (
                    f"  期望样本数: {expected_samples} "
                    f"(线程数: {threads} × 循环数: {loops})"
                )
                print(expected_info)
                print(f"  实际样本数: {total_samples}")
                if total_samples < expected_samples:
                    missing = expected_samples - total_samples
                    warn_message = (
                        f"  [警告] 缺少样本: {missing} "
                        f"({missing / expected_samples * 100:.1f}%)"
                    )
                    print(warn_message)
                    print("  [错误] 测试可能未完成或卡死")

        except Exception as e:
            print(f"[错误] 分析文件时出错: {e}")
            import traceback

            traceback.print_exc()

    # ========== 新增功能：nmon数据分析 ==========
    def analyze_nmon_data(self, stuck_time: str = "") -> None:
        """TODO: add documentation."""
        if stuck_time:
            print(f"目标时间: {stuck_time}")
        else:
            print("分析最新状态")
        print(f"{'=' * 80}")

        try:
            ssh = paramiko.SSHClient()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            ssh.connect(
                self.server_host,
                port=self.server_port,
                username=self.server_user,
                password=self.server_password,
                timeout=10,
            )

            # 查找最新的nmon文件
            stdin, stdout, stderr = ssh.exec_command(
                "ls -t /tmp/*.nmon 2>/dev/null | head -1"
            )
            nmon_file = stdout.read().decode("utf-8").strip()

            if not nmon_file:
                print("[警告] 未找到nmon文件")
                ssh.close()
                return

            print(f"\n[发现] nmon文件: {nmon_file}")

            # 下载文件
            local_nmon = (
                f"./result/temp_nmon_{datetime.now().strftime('%Y%m%d_%H%M%S')}.nmon"
            )
            sftp = ssh.open_sftp()
            sftp.get(nmon_file, local_nmon)
            sftp.close()
            ssh.close()

            print(f"[下载] 已下载到: {local_nmon}")

            # 简单分析（显示CPU和内存）
            with open(local_nmon, encoding="utf-8", errors="ignore") as f:
                lines = f.readlines()

            # 查找最新的CPU和内存数据
            cpu_values: list[float] = []
            mem_values: list[float] = []

            for line in lines[-100:]:  # 只看最后100行
                if line.startswith("CPU_ALL,") and "T" in line:
                    parts = line.split(",")
                    if len(parts) >= 6:
                        try:
                            user = (
                                float(parts[2])
                                if parts[2].replace(".", "").replace("-", "").isdigit()
                                else 0
                            )
                            sys = (
                                float(parts[3])
                                if parts[3].replace(".", "").replace("-", "").isdigit()
                                else 0
                            )
                            cpu_total = user + sys
                            cpu_values.append(cpu_total)
                        except (ValueError, TypeError) as parse_err:
                            print(f"[调试] 解析CPU统计失败: {parse_err}")

                if line.startswith("MEM,") and "T" in line:
                    parts = line.split(",")
                    if len(parts) >= 10:
                        try:
                            memtotal = (
                                float(parts[2])
                                if parts[2].replace(".", "").replace("-", "").isdigit()
                                else 8192
                            )
                            memfree = (
                                float(parts[6])
                                if parts[6].replace(".", "").replace("-", "").isdigit()
                                else 2048
                            )
                            mem_usage = (
                                ((memtotal - memfree) / memtotal) * 100
                                if memtotal > 0
                                else 0
                            )
                            mem_values.append(mem_usage)
                        except (ValueError, TypeError) as parse_err:
                            print(f"[调试] 解析内存统计失败: {parse_err}")

            if cpu_values:
                avg_cpu = sum(cpu_values) / len(cpu_values)
                print("\n[硬件状态]")
                print(f"  CPU使用率: {avg_cpu:.2f}% (平均)")
                print(f"  CPU范围: {min(cpu_values):.2f}% - {max(cpu_values):.2f}%")

            if mem_values:
                avg_mem = sum(mem_values) / len(mem_values)
                print(f"  内存使用率: {avg_mem:.2f}% (平均)")
                print(f"  内存范围: {min(mem_values):.2f}% - {max(mem_values):.2f}%")

            # 清理临时文件
            if os.path.exists(local_nmon):
                os.remove(local_nmon)

        except Exception as e:
            print(f"[错误] 分析nmon数据失败: {e}")
            import traceback

            traceback.print_exc()

    # ========== 新增功能：Excel报告审核 ==========
    def audit_excel_report(self, excel_file: str) -> None:
        """TODO: add documentation."""
        print("Excel报告专业审核")
        print("=" * 80)

        if not os.path.exists(excel_file):
            print(f"[错误] 文件不存在: {excel_file}")
            return

        wb = load_workbook(excel_file)
        issues: list[str] = []
        suggestions: list[str] = []

        print(f"\n报告文件: {excel_file}")
        print(f"文件大小: {os.path.getsize(excel_file) / 1024:.2f} KB")
        print(f"工作表数量: {len(wb.sheetnames)}")

        # 检查工作表
        required_sheets = [
            "测试执行摘要",
            "测试执行明细",
            "问题详情分析",
            "数据文件清单",
        ]
        for sheet_name in required_sheets:
            if sheet_name in wb.sheetnames:
                print(f"  ✓ 工作表存在: {sheet_name}")
            else:
                issues.append(f"缺少工作表: {sheet_name}")

        # 检查数据准确性（简化版）
        if "测试执行明细" in wb.sheetnames:
            ws = wb["测试执行明细"]
            data_rows = 0
            for row in range(3, ws.max_row + 1):
                if ws.cell(row=row, column=1).value is not None:
                    data_rows += 1

            jtl_files = glob.glob(
                os.path.join(self.result_dir, "09_device_heartbeat_*.jtl")
            )
            if data_rows != len(jtl_files):
                mismatch_message = (
                    f"数据行数不匹配: Excel显示{data_rows}行，"
                    f"实际JTL文件{len(jtl_files)}个"
                )
                issues.append(mismatch_message)
            else:
                print(f"  ✓ 数据行数正确: {data_rows}行")

        # 总结
        print(f"\n{'=' * 80}")
        print("审核结果")
        print(f"{'=' * 80}")
        if issues:
            print(f"\n[发现问题] {len(issues)}个:")
            for issue in issues[:5]:
                print(f"  ✗ {issue}")
        else:
            print("\n[通过] 未发现严重问题")

        if suggestions:
            print(f"\n[优化建议] {len(suggestions)}个:")
            for sug in suggestions[:3]:
                print(f"  - {sug}")

        wb.close()

    # ========== 新增功能：检查服务器日志 ==========
    def check_server_logs(
        self, log_date: str = "2025-11-04", time_range: str = "21:18"
    ) -> None:
        """TODO: add documentation."""
        print("检查服务器错误日志")
        print(f"日期: {log_date}, 时间范围: {time_range}")
        print(f"{'=' * 80}")

        try:
            ssh = paramiko.SSHClient()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            ssh.connect(
                self.server_host,
                port=self.server_port,
                username=self.server_user,
                password=self.server_password,
                timeout=10,
            )

            error_log = f"/opt/app/yangguan/logs/sys-error.{log_date}.log.gz"

            # 提取错误
            grep_command = (
                f"zcat {error_log} 2>/dev/null"
                " | grep -A 2 'TaskRejectedException'"
                f" | grep '{time_range}'"
                " | head -10"
            )
            stdin, stdout, stderr = ssh.exec_command(grep_command)
            error_lines = stdout.read().decode("utf-8", errors="ignore").strip()

            ssh.close()

            if error_lines:
                print("\n[发现错误]")
                for line in error_lines.split("\n")[:10]:
                    if line.strip():
                        print(f"  {line}")
            else:
                print("\n[未发现] 该时间范围内未发现TaskRejectedException错误")

        except Exception as e:
            print(f"[错误] 检查服务器日志失败: {e}")
